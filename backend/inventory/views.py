import os
import shutil
from rest_framework import serializers, viewsets, filters, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.conf import settings
from django.http import FileResponse, Http404, HttpResponse # Combinamos HttpResponse aquí
from django.db import connection, transaction, models
from django.contrib.contenttypes.models import ContentType

# Importaciones de Modelos y Serializadores (Se mantienen al final)
from .models import (
    TipoEvento, Bodega, Cliente, Manteleria, Cubierto, Loza, Cristaleria, Silla, Mesa, SalaLounge, 
    Periquera, Carpa, PistaTarima, Extra, Evento, EventoMobiliario, Degustacion, DegustacionMobiliario, Product, Notification, HomeSection, HomeSectionImage
)
from .serializers import (
    TipoEventoSerializer, BodegaSerializer, ClienteSerializer, ManteleriaSerializer, CubiertoSerializer, 
    LozaSerializer, CristaleriaSerializer, SillaSerializer, MesaSerializer, SalaLoungeSerializer, 
    PeriqueraSerializer, CarpaSerializer, PistaTarimaSerializer, ExtraSerializer, EventoSerializer, DegustacionSerializer,
    ProductSerializer, CalendarActivitySerializer, NotificationSerializer, HomeSectionSerializer, HomeSectionImageSerializer
)

# 💡 Importación ÚNICA Y CORRECTA de datetime
from datetime import datetime, timedelta 

import unidecode
from io import BytesIO
from openpyxl import Workbook
from reportlab.pdfgen import canvas # Asegúrate de tener estas importaciones si las usas abajo
from reportlab.lib.pagesizes import letter 
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
import openpyxl 

from openpyxl.styles import Font, Alignment

class TipoEventoViewSet(viewsets.ModelViewSet):
    queryset = TipoEvento.objects.all()
    serializer_class = TipoEventoSerializer
    permission_classes = [IsAuthenticated]

class BodegaViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all()
    serializer_class = BodegaSerializer
    permission_classes = [IsAuthenticated]

class MantenimientoMixin:
    @action(detail=True, methods=['post'])
    @transaction.atomic
    def mantenimiento(self, request, pk=None):
        item = self.get_object()
        cantidad_a_mantenimiento = request.data.get('cantidad', 0)

        if not isinstance(cantidad_a_mantenimiento, int) or cantidad_a_mantenimiento <= 0:
            return Response({'error': 'La cantidad debe ser un número positivo.'}, status=status.HTTP_400_BAD_REQUEST)

        if cantidad_a_mantenimiento > item.cantidad:
            return Response({'error': 'No hay suficiente stock disponible para enviar a mantenimiento.'}, status=status.HTTP_400_BAD_REQUEST)

        item.cantidad -= cantidad_a_mantenimiento
        item.cantidad_en_mantenimiento += cantidad_a_mantenimiento
        item.save()

        # Crear notificación
        message = f"Han ingresado al mantenimiento {cantidad_a_mantenimiento} {item.producto}."
        Notification.objects.create(message=message)

        return Response({'status': 'success', 'message': f'{cantidad_a_mantenimiento} unidades enviadas a mantenimiento.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def reintegrar(self, request, pk=None):
        item = self.get_object()
        cantidad_a_reintegrar = request.data.get('cantidad', 0)

        if not isinstance(cantidad_a_reintegrar, int) or cantidad_a_reintegrar <= 0:
            return Response({'error': 'La cantidad debe ser un número positivo.'}, status=status.HTTP_400_BAD_REQUEST)

        if cantidad_a_reintegrar > item.cantidad_en_mantenimiento:
            return Response({'error': 'La cantidad a reintegrar excede la que está en mantenimiento.'}, status=status.HTTP_400_BAD_REQUEST)

        item.cantidad_en_mantenimiento -= cantidad_a_reintegrar
        item.cantidad += cantidad_a_reintegrar
        item.save()

        # Crear notificación
        message = f"Han salido del mantenimiento {cantidad_a_reintegrar} {item.producto}."
        Notification.objects.create(message=message)

        return Response({'status': 'success', 'message': f'{cantidad_a_reintegrar} unidades reintegradas al stock.'}, status=status.HTTP_200_OK)


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [IsAuthenticated]

class ManteleriaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Manteleria.objects.all().order_by('-created_at')
    serializer_class = ManteleriaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class CubiertoViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Cubierto.objects.all().order_by('-created_at')
    serializer_class = CubiertoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class LozaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Loza.objects.all().order_by('-created_at')
    serializer_class = LozaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class CristaleriaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Cristaleria.objects.all().order_by('-created_at')
    serializer_class = CristaleriaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class SillaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Silla.objects.all().order_by('-created_at')
    serializer_class = SillaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class MesaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Mesa.objects.all().order_by('-created_at')
    serializer_class = MesaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class SalaLoungeViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = SalaLounge.objects.all().order_by('-created_at')
    serializer_class = SalaLoungeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class PeriqueraViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Periquera.objects.all().order_by('-created_at')
    serializer_class = PeriqueraSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class CarpaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Carpa.objects.all().order_by('-created_at')
    serializer_class = CarpaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class PistaTarimaViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = PistaTarima.objects.all().order_by('-created_at')
    serializer_class = PistaTarimaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']

class ExtraViewSet(MantenimientoMixin, viewsets.ModelViewSet):
    queryset = Extra.objects.all().order_by('-created_at')
    serializer_class = ExtraSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['producto', 'descripcion']


# --- Vistas para Eventos con lógica de negocio ---

class EventoViewSet(viewsets.ModelViewSet):
    queryset = Evento.objects.all().order_by('-created_at')
    serializer_class = EventoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return super().get_queryset().prefetch_related('mobiliario_asignado__content_object')

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        mobiliario_data = serializer.validated_data.pop('mobiliario', [])

        # 1. Validar stock
        for item in mobiliario_data:
            content_type = ContentType.objects.get_for_id(item['content_type_id'])
            model_class = content_type.model_class()
            try:
                obj = model_class.objects.get(id=item['object_id'])
                if obj.cantidad < item['cantidad']:
                    return Response(
                        {'error': f"No hay suficiente stock para {obj.producto}. Disponible: {obj.cantidad}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except model_class.DoesNotExist:
                return Response(
                    {'error': f"El item de mobiliario con id {item['object_id']} no existe."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # 2. Crear evento y reducir stock
        evento = serializer.save()
        for item in mobiliario_data:
            content_type = ContentType.objects.get_for_id(item['content_type_id'])
            model_class = content_type.model_class()
            obj = model_class.objects.get(id=item['object_id'])
            obj.cantidad -= item['cantidad']
            obj.save()

            EventoMobiliario.objects.create(
                evento=evento,
                content_type=content_type,
                object_id=obj.id,
                cantidad=item['cantidad']
            )

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        mobiliario_data = serializer.validated_data.pop('mobiliario', None)

        # Si se está actualizando el mobiliario
        if mobiliario_data is not None:
            # 1. Devolver inventario antiguo
            for item_asignado in instance.mobiliario_asignado.all():
                item_asignado.content_object.cantidad += item_asignado.cantidad
                item_asignado.content_object.save()
            instance.mobiliario_asignado.all().delete()

            # 2. Validar y asignar nuevo inventario
            for item in mobiliario_data:
                content_type = ContentType.objects.get_for_id(item['content_type_id'])
                model_class = content_type.model_class()
                obj = model_class.objects.get(id=item['object_id'])
                if obj.cantidad < item['cantidad']:
                    raise serializers.ValidationError(f"No hay suficiente stock para {obj.producto}. Disponible: {obj.cantidad}")
                
                obj.cantidad -= item['cantidad']
                obj.save()
                EventoMobiliario.objects.create(
                    evento=instance,
                    content_type=content_type,
                    object_id=obj.id,
                    cantidad=item['cantidad']
                )

        self.perform_update(serializer)
        return Response(serializer.data)


# Vista para obtener los tipos de contenido de mobiliario
class ContentTypeViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = ContentType.objects.filter(
        app_label='inventory',
        model__in=['manteleria', 'cubierto', 'loza', 'cristaleria', 'silla', 'mesa', 'salalounge', 'periquera', 'carpa', 'pistatarima', 'extra']
    ).order_by('model')
    serializer_class = serializers.Serializer # Un serializer genérico es suficiente

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        # Devolver tanto el nombre del modelo (para la lógica) como el verbose_name (para mostrar)
        data = [{'id': ct.id, 'model': ct.model, 'name': ct.model_class()._meta.verbose_name} for ct in queryset]
        return Response(data)


class DegustacionViewSet(viewsets.ModelViewSet):
    queryset = Degustacion.objects.all().order_by('-created_at')
    serializer_class = DegustacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('mobiliario_asignado__content_object')
        
        # Filtering by year and month
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        
        if year:
            queryset = queryset.filter(fecha_degustacion__year=year)
        if month:
            queryset = queryset.filter(fecha_degustacion__month=month)
            
        return queryset

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        mobiliario_data = serializer.validated_data.pop('mobiliario', [])

        for item in mobiliario_data:
            content_type = ContentType.objects.get_for_id(item['content_type_id'])
            model_class = content_type.model_class()
            try:
                obj = model_class.objects.get(id=item['object_id'])
                if obj.cantidad < item['cantidad']:
                    return Response(
                        {'error': f"No hay suficiente stock para {obj.producto}. Disponible: {obj.cantidad}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except model_class.DoesNotExist:
                return Response(
                    {'error': f"El item de mobiliario con id {item['object_id']} no existe."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        degustacion = serializer.save()
        for item in mobiliario_data:
            content_type = ContentType.objects.get_for_id(item['content_type_id'])
            model_class = content_type.model_class()
            obj = model_class.objects.get(id=item['object_id'])
            obj.cantidad -= item['cantidad']
            obj.save()

            DegustacionMobiliario.objects.create(
                degustacion=degustacion,
                content_type=content_type,
                object_id=obj.id,
                cantidad=item['cantidad']
            )

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        mobiliario_data = serializer.validated_data.pop('mobiliario', None)

        if mobiliario_data is not None:
            for item_asignado in instance.mobiliario_asignado.all():
                item_asignado.content_object.cantidad += item_asignado.cantidad
                item_asignado.content_object.save()
            instance.mobiliario_asignado.all().delete()

            for item in mobiliario_data:
                content_type = ContentType.objects.get_for_id(item['content_type_id'])
                model_class = content_type.model_class()
                obj = model_class.objects.get(id=item['object_id'])
                if obj.cantidad < item['cantidad']:
                    raise serializers.ValidationError(f"No hay suficiente stock para {obj.producto}. Disponible: {obj.cantidad}")
                
                obj.cantidad -= item['cantidad']
                obj.save()
                DegustacionMobiliario.objects.create(
                    degustacion=instance,
                    content_type=content_type,
                    object_id=obj.id,
                    cantidad=item['cantidad']
                )

        self.perform_update(serializer)
        return Response(serializer.data)


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description', 'colors']


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all().order_by('-created_at')
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def mark_all_as_read(self, request):
        Notification.objects.filter(is_read=False).update(is_read=True)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'])
    def delete_all(self, request):
        Notification.objects.all().delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CalendarDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        eventos = Evento.objects.all()
        degustaciones = Degustacion.objects.all()

        activities = []

        for evento in eventos:
            start_datetime = datetime.combine(evento.fecha_inicio, evento.hora_inicio)
            activities.append({
                'title': evento.nombre,
                'start': start_datetime,
                'end': start_datetime + timedelta(hours=2),  # Asumir duración de 2 horas
                'type': 'Evento',
                'details': {
                    'Tipo de Evento': evento.tipo_evento.nombre if evento.tipo_evento else 'No especificado',
                    'Hora del Evento': evento.hora_inicio.strftime('%H:%M'),
                    'Cantidad de Personas': evento.cantidad_personas,
                    'Ubicación': evento.lugar
                }
            })

        for degustacion in degustaciones:
            start_datetime = datetime.combine(degustacion.fecha_degustacion, degustacion.hora_degustacion)
            activities.append({
                'title': degustacion.nombre,
                'start': start_datetime,
                'end': start_datetime + timedelta(hours=1),  # Asumir duración de 1 hora
                'type': 'Degustación',
                'details': {
                    'Nombre de la Degustación': degustacion.nombre,
                    'Hora': degustacion.hora_degustacion.strftime('%H:%M'),
                    'Cantidad de Personas': degustacion.cantidad_personas
                }
            })

        serializer = CalendarActivitySerializer(activities, many=True)
        return Response(serializer.data)


from django.http import JsonResponse, HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment

class LowStockInventoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """
        Returns a list of all inventory items with stock below 25 units.
        """
        # Define the inventory models to check
        inventory_models = [
            Manteleria, Cubierto, Loza, Cristaleria, Silla,
            Mesa, SalaLounge, Periquera, Carpa, PistaTarima, Extra
        ]

        low_stock_items = []

        for model in inventory_models:
            # Get all items with stock below 25
            items = model.objects.filter(cantidad__lt=25).select_related('bodega')

            for item in items:
                low_stock_items.append({
                    'id': item.id,
                    'categoria': model._meta.verbose_name_plural.title(),
                    'nombre': item.producto,
                    'descripcion': item.descripcion,
                    'cantidad_actual': item.cantidad,
                    'stock_minimo': 25,  # Default minimum stock level
                    'bodega_id': item.bodega.id if item.bodega else None,
                    'bodega_nombre': item.bodega.nombre if item.bodega else 'No especificada',
                    'tipo': model.__name__.lower()
                })

        # Sort by category and then by current quantity (ascending)
        low_stock_items.sort(key=lambda x: (x['categoria'], x['cantidad_actual']))

        return Response(low_stock_items)


class MaintenanceReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """
        Returns a list of all furniture items currently in maintenance or that 
        were in maintenance activity (entry/exit) in the last 30 days.
        """
        # Define the inventory models to check (furniture/mobiliary items)
        inventory_models = [
            Manteleria, Cubierto, Loza, Cristaleria, Silla,
            Mesa, SalaLounge, Periquera, Carpa, PistaTarima, Extra
        ]

        maintenance_items = []
        
        # Get date range for the last 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        # 1. Get items currently in maintenance
        for model in inventory_models:
            items = model.objects.filter(cantidad_en_mantenimiento__gt=0).select_related('bodega')
            
            for item in items:
                maintenance_items.append({
                    'id': item.id,
                    'categoria': model._meta.verbose_name_plural.title(),
                    'nombre': item.producto,
                    'descripcion': item.descripcion,
                    'cantidad_en_mantenimiento': item.cantidad_en_mantenimiento,
                    'cantidad_disponible': item.cantidad,
                    'bodega_id': item.bodega.id if item.bodega else None,
                    'bodega_nombre': item.bodega.nombre if item.bodega else 'No especificada',
                    'estado': 'En Mantenimiento',
                    'fecha': item.updated_at.isoformat() if item.updated_at else None,
                    'tipo': model.__name__.lower()
                })

        # 2. Get maintenance activity from notifications in the last 30 days
        maintenance_notifications = Notification.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).filter(
            models.Q(message__icontains='ingresado al mantenimiento') | 
            models.Q(message__icontains='salido del mantenimiento')
        ).order_by('-created_at')

        # Parse notifications to extract maintenance history
        for notification in maintenance_notifications:
            message = notification.message
            
            # Determine if it's entry or exit from maintenance
            if 'ingresado al mantenimiento' in message.lower():
                estado = 'Ingresó a Mantenimiento'
            elif 'salido del mantenimiento' in message.lower():
                estado = 'Salió de Mantenimiento'
            else:
                continue
            
            # Try to extract quantity and product name from message
            # Message format: "Han ingresado/salido del mantenimiento {cantidad} {producto}."
            import re
            match = re.search(r'(\d+)\s+(.+?)\.$', message)
            if match:
                cantidad = int(match.group(1))
                producto = match.group(2).strip()
                
                # Find the model that matches this product
                found_item = None
                for model in inventory_models:
                    try:
                        item = model.objects.filter(producto=producto).first()
                        if item:
                            found_item = {
                                'id': f"notif_{notification.id}",
                                'categoria': model._meta.verbose_name_plural.title(),
                                'nombre': producto,
                                'descripcion': '',
                                'cantidad_en_mantenimiento': cantidad,
                                'cantidad_disponible': item.cantidad if item else 0,
                                'bodega_id': item.bodega.id if item and item.bodega else None,
                                'bodega_nombre': item.bodega.nombre if item and item.bodega else 'No especificada',
                                'estado': estado,
                                'fecha': notification.created_at.isoformat(),
                                'tipo': model.__name__.lower()
                            }
                            break
                    except Exception as e:
                        continue
                
                if found_item:
                    # Check if this item is not already in the list (avoid duplicates with current maintenance)
                    is_duplicate = any(
                        existing['nombre'] == found_item['nombre'] and 
                        existing['estado'] == 'En Mantenimiento'
                        for existing in maintenance_items
                    )
                    if not is_duplicate:
                        maintenance_items.append(found_item)

        # Sort by date (most recent first) and then by category
        maintenance_items.sort(key=lambda x: (x.get('fecha', ''), x['categoria']), reverse=True)

        return Response(maintenance_items)


class EventAnalysisReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """
        Returns event analysis grouped by time periods (monthly, quarterly, yearly)
        with count and percentage for each period.
        """
        period_type = request.query_params.get('period', 'monthly')  # monthly, quarterly, yearly
        
        # Get all events
        eventos = Evento.objects.all().order_by('fecha_inicio')
        
        if not eventos.exists():
            return Response({
                'period_type': period_type,
                'periods': [],
                'total_events': 0
            })
        
        # Group events by period
        period_data = {}
        total_events = eventos.count()
        
        for evento in eventos:
            fecha = evento.fecha_inicio
            
            if period_type == 'monthly':
                # Format: "Enero 2024", "Febrero 2024", etc.
                period_key = fecha.strftime('%B %Y')
                # Spanish month names
                month_names = {
                    'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
                    'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
                    'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
                    'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
                }
                for eng, esp in month_names.items():
                    period_key = period_key.replace(eng, esp)
                    
            elif period_type == 'quarterly':
                # Format: "Q1 2024", "Q2 2024", etc.
                quarter = (fecha.month - 1) // 3 + 1
                period_key = f"Q{quarter} {fecha.year}"
                
            else:  # yearly
                # Format: "2024", "2025", etc.
                period_key = str(fecha.year)
            
            if period_key not in period_data:
                period_data[period_key] = {
                    'period': period_key,
                    'count': 0,
                    'events': []
                }
            
            period_data[period_key]['count'] += 1
            period_data[period_key]['events'].append({
                'id': evento.id,
                'nombre': evento.nombre,
                'fecha': evento.fecha_inicio.isoformat(),
                'responsable': evento.responsable,
                'cantidad_personas': evento.cantidad_personas
            })
        
        # Calculate percentages and format response
        periods = []
        for period_key, data in period_data.items():
            percentage = round((data['count'] / total_events) * 100, 1)
            periods.append({
                'period': data['period'],
                'count': data['count'],
                'percentage': percentage,
                'events': data['events']
            })
        
        # Sort periods chronologically
        if period_type == 'yearly':
            periods.sort(key=lambda x: x['period'])
        elif period_type == 'quarterly':
            periods.sort(key=lambda x: (x['period'].split()[1], x['period'].split()[0]))
        else:  # monthly
            # Sort by date parsing
            month_order = {
                'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
                'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
                'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
            }
            periods.sort(key=lambda x: (
                int(x['period'].split()[1]),  # year
                month_order.get(x['period'].split()[0], 0)  # month
            ))
        
        return Response({
            'period_type': period_type,
            'periods': periods,
            'total_events': total_events
        })


class InventoryUsageReportView(APIView):
    permission_classes = [IsAuthenticated]

    def generate_test_excel(self):
        """Genera un archivo Excel de prueba directamente."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Prueba"
        
        # ... (Tus cabeceras y datos de prueba) ...
        
        # 1. Crear un buffer en memoria (BytesIO)
        output = BytesIO()
        
        # 2. Guardar el workbook en el buffer
        wb.save(output)
        
        # 3. Mover el puntero al inicio del buffer
        output.seek(0)

        # 4. Generar la respuesta HTTP desde el buffer
        response = HttpResponse(
            output.read(), # Leemos el contenido binario del buffer
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Clave para la descarga: Cabecera Content-Disposition
        response['Content-Disposition'] = 'attachment; filename="reporte_de_prueba.xlsx"'
        return response
        
    
    # --- MÉTODO GET MODIFICADO ---
    def get(self, request, *args, **kwargs):
        # Parámetros para la descarga de reportes
        report_format = request.query_params.get('format')
        event_id = request.query_params.get('event_id')

        # Parámetros para el listado de eventos
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        # 🎯 CASO 1: PRUEBA DE DESCARGA (Test endpoint)
        if 'new-report-download' in request.path:
            return self.generate_test_excel()
        
        # -----------------------------------------------------

        # 🎯 CASO 2: DESCARGAR UN REPORTE DE EVENTO ESPECÍFICO (reports/download/)
        if report_format and event_id:
            try:
                evento = Evento.objects.get(pk=event_id)
                if report_format == 'pdf':
                    # Llama a tu función de generación de PDF
                    return self.generate_pdf(evento)
                elif report_format == 'excel':
                    # Llama a tu función de generación de Excel
                    return self.generate_excel(evento)
                else:
                    return Response({"error": "Formato de reporte no válido."}, status=status.HTTP_400_BAD_REQUEST)
            
            except Evento.DoesNotExist:
                return Response({"error": f"Evento con ID {event_id} no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        # -----------------------------------------------------

        # 🎯 CASO 3: LISTAR EVENTOS POR RANGO DE FECHAS (reports/)
        elif start_date_str and end_date_str:
            try:
                # Intenta convertir las cadenas de fecha
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                
            except ValueError as e:
                # 🛑 Devuelve 400 Bad Request si la conversión falla (solución al error anterior)
                print(f"ERROR AL PARSEAR FECHAS: {e}")
                return Response({
                    "error": "Formato de fecha inválido. Asegúrate de usar AAAA-MM-DD.",
                    "detalle": str(e)
                }, status=status.HTTP_400_BAD_REQUEST)

            # Si la conversión fue exitosa:
            events = Evento.objects.filter(fecha_inicio__range=[start_date, end_date])
            
            # Serializa la respuesta para enviarla al frontend
            serializer = EventoSerializer(events, many=True)
            return Response(serializer.data)

        # -----------------------------------------------------

        # CASO 4: Si no se proporcionaron parámetros suficientes/válidos
        return Response({
            "error": "Parámetros incorrectos. Se requiere 'start_date' y 'end_date' o 'format' y 'event_id'."
        }, status=status.HTTP_400_BAD_REQUEST)

    def generate_pdf(self, evento):
        # 1. PREPARAR EL NOMBRE DEL ARCHIVO SEGURO
        
        # Eliminamos acentos, tildes, y caracteres problemáticos del nombre del evento
        # para que el nombre de archivo sea compatible con la mayoría de los sistemas.
        # Esto previene errores de codificación complejos en la cabecera.
        import unidecode
        
        # Usamos unidecode para simplificar caracteres (ej: Evento Día Ñ -> Evento Dia N)
        clean_name = unidecode.unidecode(evento.nombre).replace(" ", "_").replace("/", "-")
        base_filename = f"reporte_evento_{clean_name}.pdf"

        # 2. GENERAR LA RESPUESTA HTTP Y CABECERAS (CRÍTICO)
        response = HttpResponse(content_type='application/pdf')
        
        # Codificamos la cabecera 'Content-Disposition' con latin-1, el estándar de Django
        # para asegurar que los navegadores manejen correctamente los nombres de archivo no-ASCII.
        response['Content-Disposition'] = f'attachment; filename="{base_filename}"'.encode('latin-1', errors='replace').decode('latin-1')
        
        # 3. GENERACIÓN DEL CONTENIDO DEL PDF CON REPORTLAB
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Título y detalles
        story.append(Paragraph(f"Reporte de Uso de Inventario", styles['Title']))
        story.append(Spacer(1, 12))

        # Nota: Usar 'Normal' con <b> funciona, pero si hay problemas de codificación 
        # o fuentes en ReportLab, es mejor mantener el texto simple.
        story.append(Paragraph(f"<b>Nombre del Evento:</b> {evento.nombre}", styles['Normal']))
        story.append(Paragraph(f"<b>Tipo de Evento:</b> {evento.tipo_evento.nombre if evento.tipo_evento else 'N/A'}", styles['Normal']))
        story.append(Paragraph(f"<b>Responsable:</b> {evento.responsable}", styles['Normal']))
        story.append(Paragraph(f"<b>Lugar:</b> {evento.lugar}", styles['Normal']))
        story.append(Paragraph(f"<b>Fecha:</b> {evento.fecha_inicio.strftime('%d/%m/%Y')}", styles['Normal']))
        story.append(Spacer(1, 24))

        story.append(Paragraph("Inventario Utilizado", styles['h2']))
        
        # Datos de la tabla
        mobiliario_data = [['Producto', 'Descripción', 'Cantidad']]
        for item in evento.mobiliario_asignado.all():
            mobiliario_data.append([
                item.content_object.producto,
                item.content_object.descripcion,
                item.cantidad
            ])

        table = Table(mobiliario_data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)
        story.append(table)

        doc.build(story)
        return response

    def generate_excel(self, evento):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_evento_{evento.nombre}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Inventario"

        # --- Header ---
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Reporte de Uso de Inventario"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')

        # --- Event Details ---
        details = [
            ("Nombre del Evento:", evento.nombre),
            ("Tipo de Evento:", evento.tipo_evento.nombre if evento.tipo_evento else 'N/A'),
            ("Responsable:", evento.responsable),
            ("Lugar:", evento.lugar),
            ("Fecha:", evento.fecha_inicio.strftime('%d/%m/%Y'))
        ]
        row = 3
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # --- Inventory Table ---
        table_header_row = row + 1
        headers = ["Producto", "Descripción", "Cantidad"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=table_header_row, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True)

        for item in evento.mobiliario_asignado.all():
            table_header_row += 1
            ws.cell(row=table_header_row, column=1).value = item.content_object.producto
            ws.cell(row=table_header_row, column=2).value = item.content_object.descripcion
            ws.cell(row=table_header_row, column=3).value = item.cantidad

        wb.save(response)
        return response


class BackupCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # 1. Verificar que estamos usando SQLite (o ajustar si es otro motor)
        if 'sqlite3' not in settings.DATABASES['default']['ENGINE']:
            return Response({'error': 'La función de respaldo solo está configurada para SQLite.'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        db_path = settings.DATABASES['default']['NAME']
        
        # 2. Generar un nombre de archivo dinámico
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"db_backup_{timestamp}.sqlite3"
        
        # 3. Asegurar que la conexión a la base de datos esté cerrada temporalmente (Necesario para copiar SQLite)
        # Esto es crucial para asegurar una copia consistente de SQLite
        connection.close() 

        try:
            # 4. Usar FileResponse para servir el archivo directamente
            response = FileResponse(
                open(db_path, 'rb'), 
                content_type='application/octet-stream'
            )
            
            # 5. Establecer el encabezado Content-Disposition (crucial para la descarga)
            response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
            
            return response
            
        except FileNotFoundError:
            raise Http404("Archivo de base de datos no encontrado.")
            
        finally:
            # 6. Reabrir la conexión a la base de datos después de la operación
            connection.ensure_connection()

class BackupRestoreView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        # 1. Obtener el archivo del request (clave 'backup_file' del frontend)
        uploaded_file = request.FILES.get('backup_file')
        
        if not uploaded_file:
            return Response({'error': 'No se encontró el archivo de respaldo.'}, 
                            status=status.HTTP_400_BAD_REQUEST)

        # 2. Configuración de rutas (igual que en BackupCreateView)
        db_path = settings.DATABASES['default']['NAME']
        
        # 3. Guardar el archivo subido en una ubicación temporal
        temp_dir = settings.BASE_DIR / 'temp_restore'
        os.makedirs(temp_dir, exist_ok=True)
        temp_file_path = temp_dir / 'incoming_backup.sqlite3'
        
        with open(temp_file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        # 4. CRÍTICO: Cerrar la conexión a la base de datos
        connection.close() 

        try:
            # 5. Sobrescribir el archivo de base de datos con el archivo de respaldo
            shutil.copyfile(temp_file_path, db_path)
            
            # 6. Responder éxito
            return Response({'status': 'Restauración completada exitosamente. Se recomienda recargar el sistema.'}, 
                            status=status.HTTP_200_OK)

        except Exception as e:
            # 7. Manejo de errores
            return Response({'error': f'Error al restaurar la base de datos: {str(e)}'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finally:
            # 8. Reabrir la conexión y limpiar el archivo temporal
            connection.ensure_connection()
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)


class WarehouseInventoryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """
        Returns inventory data grouped by warehouse and category.
        Shows percentage of inventory per warehouse and items by category.
        """
        # Define the inventory models to check
        inventory_models = [
            ('Manteleria', Manteleria),
            ('Cubierto', Cubierto),
            ('Loza', Loza),
            ('Cristaleria', Cristaleria),
            ('Sillas', Silla),
            ('Mesas', Mesa),
            ('Salas lounge', SalaLounge),
            ('Periqueras', Periquera),
            ('Carpas', Carpa),
            ('Pistas y tarimas', PistaTarima),
            ('Extras', Extra)
        ]

        # Get all warehouses
        bodegas = Bodega.objects.all()
        
        # Structure to hold the report data
        report_data = []
        total_inventory = 0
        
        # Calculate total inventory across all warehouses
        for category_name, model in inventory_models:
            total_inventory += model.objects.aggregate(total=models.Sum('cantidad'))['total'] or 0
        
        # Process each warehouse
        for bodega in bodegas:
            bodega_data = {
                'id': bodega.id,
                'nombre': bodega.nombre,
                'ubicacion': bodega.ubicacion,
                'total_items': 0,
                'percentage': 0,
                'categories': []
            }
            
            bodega_total = 0
            category_details = []
            
            # Calculate inventory by category for this warehouse
            for category_name, model in inventory_models:
                category_total = model.objects.filter(bodega=bodega).aggregate(
                    total=models.Sum('cantidad')
                )['total'] or 0
                
                bodega_total += category_total
                
                category_details.append({
                    'categoria': category_name,
                    'cantidad': category_total,
                    'percentage': 0  # Will calculate after we know bodega_total
                })
            
            # Calculate percentages
            bodega_data['total_items'] = bodega_total
    def get(self, request, *args, **kwargs):
        # 1. Verificar que estamos usando SQLite (o ajustar si es otro motor)
        if 'sqlite3' not in settings.DATABASES['default']['ENGINE']:
            return Response({'error': 'La función de respaldo solo está configurada para SQLite.'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        db_path = settings.DATABASES['default']['NAME']
        
        # 2. Generar un nombre de archivo dinámico
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"db_backup_{timestamp}.sqlite3"
        
        # 3. Asegurar que la conexión a la base de datos esté cerrada temporalmente (Necesario para copiar SQLite)
        # Esto es crucial para asegurar una copia consistente de SQLite
        connection.close() 

        try:
            # 4. Usar FileResponse para servir el archivo directamente
            response = FileResponse(
                open(db_path, 'rb'), 
                content_type='application/octet-stream'
            )
            
            # 5. Establecer el encabezado Content-Disposition (crucial para la descarga)
            response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
            
            return response
            
        except FileNotFoundError:
            raise Http404("Archivo de base de datos no encontrado.")
            
        finally:
            # 6. Reabrir la conexión a la base de datos después de la operación
            connection.ensure_connection()

class BackupRestoreView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        # 1. Obtener el archivo del request (clave 'backup_file' del frontend)
        uploaded_file = request.FILES.get('backup_file')
        
        if not uploaded_file:
            return Response({'error': 'No se encontró el archivo de respaldo.'}, 
                            status=status.HTTP_400_BAD_REQUEST)

        # 2. Configuración de rutas (igual que en BackupCreateView)
        db_path = settings.DATABASES['default']['NAME']
        
        # 3. Guardar el archivo subido en una ubicación temporal
        temp_dir = settings.BASE_DIR / 'temp_restore'
        os.makedirs(temp_dir, exist_ok=True)
        temp_file_path = temp_dir / 'incoming_backup.sqlite3'
        
        with open(temp_file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        # 4. CRÍTICO: Cerrar la conexión a la base de datos
        connection.close() 

        try:
            # 5. Sobrescribir el archivo de base de datos con el archivo de respaldo
            shutil.copyfile(temp_file_path, db_path)
            
            # 6. Responder éxito
            return Response({'status': 'Restauración completada exitosamente. Se recomienda recargar el sistema.'}, 
                            status=status.HTTP_200_OK)

        except Exception as e:
            # 7. Manejo de errores
            return Response({'error': f'Error al restaurar la base de datos: {str(e)}'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finally:
            # 8. Reabrir la conexión y limpiar el archivo temporal
            connection.ensure_connection()
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)


class WarehouseInventoryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        """
        Returns inventory data grouped by warehouse and category.
        Shows percentage of inventory per warehouse and items by category.
        """
        # Define the inventory models to check
        inventory_models = [
            ('Manteleria', Manteleria),
            ('Cubierto', Cubierto),
            ('Loza', Loza),
            ('Cristaleria', Cristaleria),
            ('Sillas', Silla),
            ('Mesas', Mesa),
            ('Salas lounge', SalaLounge),
            ('Periqueras', Periquera),
            ('Carpas', Carpa),
            ('Pistas y tarimas', PistaTarima),
            ('Extras', Extra)
        ]

        # Get all warehouses
        bodegas = Bodega.objects.all()
        
        # Structure to hold the report data
        report_data = []
        total_inventory = 0
        
        # Calculate total inventory across all warehouses
        for category_name, model in inventory_models:
            total_inventory += model.objects.aggregate(total=models.Sum('cantidad'))['total'] or 0
        
        # Process each warehouse
        for bodega in bodegas:
            bodega_data = {
                'id': bodega.id,
                'nombre': bodega.nombre,
                'ubicacion': bodega.ubicacion,
                'total_items': 0,
                'percentage': 0,
                'categories': []
            }
            
            bodega_total = 0
            category_details = []
            
            # Calculate inventory by category for this warehouse
            for category_name, model in inventory_models:
                category_total = model.objects.filter(bodega=bodega).aggregate(
                    total=models.Sum('cantidad')
                )['total'] or 0
                
                bodega_total += category_total
                
                category_details.append({
                    'categoria': category_name,
                    'cantidad': category_total,
                    'percentage': 0  # Will calculate after we know bodega_total
                })
            
            # Calculate percentages
            bodega_data['total_items'] = bodega_total
            if total_inventory > 0:
                bodega_data['percentage'] = round((bodega_total / total_inventory) * 100, 2)
            
            # Calculate category percentages within this warehouse
            for category in category_details:
                if bodega_total > 0:
                    category['percentage'] = round((category['cantidad'] / bodega_total) * 100, 2)
            
            bodega_data['categories'] = category_details
            report_data.append(bodega_data)
        
        return Response({
            'total_inventory': total_inventory,
            'warehouses': report_data
        })

class HomeSectionViewSet(viewsets.ModelViewSet):
    queryset = HomeSection.objects.all()
    serializer_class = HomeSectionSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'name'

    @action(detail=True, methods=['post'])
    def upload_image(self, request, name=None):
        section = self.get_object()
        image = request.FILES.get('image')
        caption = request.data.get('caption', '')
        
        if not image:
            return Response({'error': 'No image provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Optional: Clear existing images if you only want one per section (or specific logic)
        # section.images.all().delete() 

        HomeSectionImage.objects.create(section=section, image=image, caption=caption)
        
        # Return updated section data
        serializer = self.get_serializer(section)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def delete_image(self, request, name=None):
        image_id = request.data.get('image_id')
        if not image_id:
             return Response({'error': 'No image_id provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            image = HomeSectionImage.objects.get(id=image_id, section__name=name)
            image.delete()
            return Response({'status': 'Image deleted'})
        except HomeSectionImage.DoesNotExist:
            return Response({'error': 'Image not found'}, status=status.HTTP_404_NOT_FOUND)